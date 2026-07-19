"""Smoke test: modern threaded comments in Excel (.xlsx).

add -> reply -> resolve/reopen -> delete via _xlsx_comments, asserting the threaded part,
the persons part, the regenerated legacy shadow (comments + VML + legacyDrawing), threading,
the resolved flag, byte-stability of unrelated parts, and that openpyxl still opens each step.

Run: python tests/smoke_xlsx_comments.py
"""
import os
import sys
import tempfile
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "scripts"))
import _xlsx_comments as xc   # noqa: E402
import _errors                # noqa: E402
from _ooxml_zip import patch_parts             # noqa: E402
from lxml import etree                         # noqa: E402
from openpyxl import Workbook, load_workbook   # noqa: E402

XNS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RNS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_VML_NOTE = (
    '<xml xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office"'
    ' xmlns:x="urn:schemas-microsoft-com:office:excel">'
    '<o:shapelayout v:ext="edit"><o:idmap v:ext="edit" data="1"/></o:shapelayout>'
    '<v:shapetype id="_x0000_t202" coordsize="21600,21600" o:spt="202" path="m,l,21600r21600,l21600,xe">'
    '<v:stroke joinstyle="miter"/><v:path gradientshapeok="t" o:connecttype="rect"/></v:shapetype>'
    '{shapes}</xml>')
_NOTE_SHAPE = (
    '<v:shape id="_x0000_s{sid}" type="#_x0000_t202" style="position:absolute;visibility:hidden"'
    ' fillcolor="#ffffe1" o:insetmode="auto"><v:fill color2="#ffffe1"/>'
    '<v:shadow on="t" color="black" obscured="t"/><v:path o:connecttype="none"/>'
    '<x:ClientData ObjectType="{otype}"><x:MoveWithCells/><x:SizeWithCells/>'
    '<x:AutoFill>False</x:AutoFill>{rowcol}'
    '</x:ClientData></v:shape>')


def _rowcol(row, col):
    return f"<x:Row>{row}</x:Row><x:Column>{col}</x:Column>"


def _add_classic_note(path, ref="D4", author="Deval", text="Classic note text",
                      with_row_col=True):
    """Inject a classic (non-threaded) note the way real Excel stores one: a comments
    part with a human author, a VML note shape, both sheet rels, and <legacyDrawing>.
    `ref` is written into the XML VERBATIM (so 'd4' tests non-canonical spellings);
    with_row_col=False emits a shape without <x:Row>/<x:Column> (foreign generators)."""
    col0 = ord(ref[0].upper()) - 65
    row0 = int(ref[1:]) - 1

    def mut(ps):
        comments = (
            f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            f'<comments xmlns="{XNS}"><authors><author>{author}</author></authors>'
            f'<commentList><comment ref="{ref}" authorId="0" shapeId="0">'
            f'<text><t>{text}</t></text></comment></commentList></comments>')
        ps.set_bytes("xl/comments1.xml", comments.encode("utf-8"))
        ps.ensure_content_type(
            "/xl/comments1.xml",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml")
        ps.ensure_default("vml", "application/vnd.openxmlformats-officedocument.vmlDrawing")
        shape = _NOTE_SHAPE.format(sid=1025, otype="Note",
                                   rowcol=_rowcol(row0, col0) if with_row_col else "")
        ps.set_bytes("xl/drawings/vmlDrawing1.vml", _VML_NOTE.format(shapes=shape).encode("utf-8"))
        rels = ps.rels_for("xl/worksheets/sheet1.xml")
        rels.add_rel(RNS + "/comments", "../comments1.xml")
        rid_v = rels.add_rel(RNS + "/vmlDrawing", "../drawings/vmlDrawing1.vml")
        sroot = ps.get_xml("xl/worksheets/sheet1.xml")
        ld = etree.SubElement(sroot, f"{{{XNS}}}legacyDrawing", nsmap={"r": RNS})
        ld.set(f"{{{RNS}}}id", rid_v)
        ps.set_xml("xl/worksheets/sheet1.xml", sroot)
    patch_parts(path, mut)


def _base(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"], ws["B1"] = "Parameter", "Value"
    ws["A2"], ws["B2"] = "North region", "$1.2M"
    wb.save(path)


def _opens(path):
    try:
        load_workbook(path)
        return True
    except Exception as e:
        return f"{type(e).__name__}: {e}"


def _parts(path):
    with zipfile.ZipFile(path) as z:
        return {i.filename: z.read(i.filename) for i in z.infolist()}


def main() -> int:
    fails = []
    tmp = tempfile.mkdtemp(prefix="xlsx_comments_")
    f = os.path.join(tmp, "t.xlsx")
    _base(f)
    before = _parts(f)

    try:
        xc.add_comment(f, {"sheet": "Nope", "cell": "B2"}, "x", author="A")
        fails.append("bad sheet: expected AnchorNotFound")
    except _errors.AnchorNotFound:
        pass
    except Exception as e:
        fails.append(f"bad sheet: wrong error {type(e).__name__}")

    # --- add ---
    cid = xc.add_comment(f, {"sheet": "Data", "cell": "B2"},
                         "Please confirm this figure matches the latest forecast.", author="Alex Morgan")
    recs = xc.list_comments(f)
    if len(recs) != 1:
        fails.append(f"add: expected 1 comment, got {len(recs)}")
    else:
        r = recs[0]
        if r["anchor"]["cell"] != "B2" or r["location"] != "Data!B2":
            fails.append(f"add: bad anchor/location {r['anchor']} {r['location']}")
        if r["author"] != "Alex Morgan":
            fails.append(f"add: author {r['author']!r}")
        if "$1.2M" not in r["context"]:
            fails.append(f"add: context missing cell value: {r['context']!r}")
    with zipfile.ZipFile(f) as z:
        names = set(z.namelist())
        for p in ("xl/threadedComments/threadedComment1.xml", "xl/persons/person.xml", "xl/comments1.xml"):
            if p not in names:
                fails.append(f"add: missing part {p}")
        if not any(n.endswith(".vml") for n in names):
            fails.append("add: no VML drawing written")
        ct = z.read("[Content_Types].xml").decode()
        for frag in ("ms-excel.threadedcomments", "ms-excel.person", "spreadsheetml.comments", "vmlDrawing"):
            if frag not in ct:
                fails.append(f"add: content-type missing {frag}")
        srels = z.read("xl/worksheets/_rels/sheet1.xml.rels").decode()
        for frag in ("relationships/threadedComment", "relationships/comments", "relationships/vmlDrawing"):
            if frag not in srels:
                fails.append(f"add: sheet rel missing {frag}")
        sheet_xml = z.read("xl/worksheets/sheet1.xml").decode()
        if "legacyDrawing" not in sheet_xml:
            fails.append("add: legacyDrawing not added to sheet")
        if "r:id=" not in sheet_xml:
            fails.append("add: legacyDrawing missing r:id (relationships prefix)")
        if "ns0:id=" in sheet_xml:
            fails.append("add: legacyDrawing serialised with ns0:id instead of r:id (#1 regression)")
    if (o := _opens(f)) is not True:
        fails.append(f"add: won't open: {o}")
    changed = sorted(n for n in before if before[n] != _parts(f).get(n))
    allowed = {"[Content_Types].xml", "xl/worksheets/sheet1.xml", "xl/_rels/workbook.xml.rels"}
    if not set(changed).issubset(allowed):
        fails.append(f"add: changed unexpected existing parts: {set(changed) - allowed}")
    if "[Content_Types].xml" not in changed or "xl/worksheets/sheet1.xml" not in changed:
        fails.append(f"add: expected content-types + sheet to change, got {changed}")

    # --- reply ---
    rid = xc.reply(f, cid, "Confirmed - it matches the final forecast.", author="Sam Lee")
    recs = {r["id"]: r for r in xc.list_comments(f)}
    if len(recs) != 2:
        fails.append(f"reply: expected 2, got {len(recs)}")
    if rid in recs:
        if recs[rid]["parent_id"] != cid or recs[rid]["thread_id"] != cid or not recs[rid]["is_reply"]:
            fails.append(f"reply: bad threading {recs[rid]}")
    else:
        fails.append("reply: reply id not listed")
    if (o := _opens(f)) is not True:
        fails.append(f"reply: won't open: {o}")

    # --- resolve / reopen ---
    xc.set_status(f, cid, True)
    if not all(r["resolved"] for r in xc.list_comments(f)):
        fails.append("resolve: not marked resolved")
    xc.set_status(f, rid, False)
    if any(r["resolved"] for r in xc.list_comments(f)):
        fails.append("reopen: still resolved")

    # --- delete reply then root ---
    xc.delete(f, rid)
    if [r["id"] for r in xc.list_comments(f)] != [cid]:
        fails.append("delete reply: wrong remaining set")
    xc.delete(f, cid)
    if xc.list_comments(f):
        fails.append("delete root: comments remain")
    if (o := _opens(f)) is not True:
        fails.append(f"final: won't open: {o}")

    # --- multi-comment regression: more than one comment per sheet must keep shapeId="0" on every
    #     legacy note. A per-comment shapeId (0,1,2..) points at a nonexistent VML shape, so Excel
    #     drops the whole comments part on open. Also covers comments on a non-first sheet. ---
    import re
    f2 = os.path.join(tmp, "multi.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    s2 = wb.create_sheet("Second")
    for i in range(1, 6):
        ws[f"B{i}"] = f"v{i}"
        s2[f"B{i}"] = f"w{i}"
    wb.save(f2)
    for i in range(1, 6):
        mcid = xc.add_comment(f2, {"sheet": "Data", "cell": f"B{i}"}, f"C{i}", author="Alex Morgan")
        xc.reply(f2, mcid, f"R{i}", author="Sam Lee")
    xc.add_comment(f2, {"sheet": "Second", "cell": "B1"}, "on the second sheet", author="Alex Morgan")

    if len(xc.list_comments(f2)) != 11:   # 5 roots + 5 replies on Data, 1 on Second
        fails.append(f"multi: expected 11 comments, got {len(xc.list_comments(f2))}")
    if (o := _opens(f2)) is not True:
        fails.append(f"multi: won't open: {o}")
    with zipfile.ZipFile(f2) as z:
        cparts = [n for n in z.namelist() if re.match(r"xl/comments\d+\.xml$", n)]
        if "xl/comments1.xml" not in cparts or "xl/comments2.xml" not in cparts:
            fails.append(f"multi: expected per-sheet comments1.xml + comments2.xml, got {sorted(cparts)}")
        for cpart in cparts:
            body = z.read(cpart).decode()
            shape_ids = re.findall(r'shapeId="([^"]*)"', body)
            if any(s != "0" for s in shape_ids):
                fails.append(f"multi: {cpart} has a non-zero shapeId {shape_ids}; Excel drops the "
                             f"comments unless every legacy note has shapeId=0")
            if body.count("<comment ") != body.count("<author>"):
                fails.append(f"multi: {cpart} comment/author count mismatch")

    # --- foreign relationship spelling: adding a comment must REUSE the sheet's existing
    #     comments/vml rel, not append a duplicate (duplicate rels make Excel drop all comments) ---
    f3 = os.path.join(tmp, "foreign.xlsx")
    _base(f3)
    xc.add_comment(f3, {"sheet": "Data", "cell": "B2"}, "one", author="Alex Morgan")
    rp = "xl/worksheets/_rels/sheet1.xml.rels"
    with zipfile.ZipFile(f3) as z:
        infos = z.infolist()
        data = {i.filename: z.read(i.filename) for i in infos}
    data[rp] = (data[rp].decode()
                .replace('Target="../comments1.xml"', 'Target="/xl/comments1.xml"')
                .replace('Target="../drawings/vmlDrawing1.vml"', 'Target="/xl/drawings/vmlDrawing1.vml"')
                .encode())
    with zipfile.ZipFile(f3, "w", zipfile.ZIP_DEFLATED) as zo:
        for i in infos:
            zo.writestr(i, data[i.filename])
    xc.add_comment(f3, {"sheet": "Data", "cell": "B3"}, "two", author="Alex Morgan")
    with zipfile.ZipFile(f3) as z:
        srels = z.read(rp).decode()
    for frag, name in (("relationships/comments", "comments"), ("relationships/vmlDrawing", "vmlDrawing")):
        if srels.count(frag) != 1:
            fails.append(f"foreign-rels: expected 1 {name} rel, got {srels.count(frag)} "
                         f"(a duplicate rel makes Excel drop the comments)")
    if (o := _opens(f3)) is not True:
        fails.append(f"foreign-rels: won't open: {o}")

    # --- v0.3.0 (#3): invalid cell references rejected upfront; refs canonicalized ---
    f4 = os.path.join(tmp, "refs.xlsx")
    _base(f4)
    for bad in ("A0", "B0", "XFE1", "A1048577"):
        try:
            xc.add_comment(f4, {"sheet": "Data", "cell": bad}, "x", author="A")
            fails.append(f"ref {bad}: expected AnchorNotFound")
        except _errors.AnchorNotFound:
            pass
        except Exception as e:  # noqa: BLE001
            fails.append(f"ref {bad}: wrong error {type(e).__name__}: {e}")
    xc.add_comment(f4, {"sheet": "Data", "cell": "a01"}, "canonical?", author="Alex Morgan")
    recs = xc.list_comments(f4)
    if len(recs) != 1 or recs[0]["anchor"]["cell"] != "A1":
        fails.append(f"ref a01: expected canonical A1, got {[r['anchor'] for r in recs]}")

    # --- v0.3.0: classic-note COEXISTENCE ------------------------------------------
    # (a) adding a thread to a sheet holding a classic note now WORKS (old: refused),
    #     and the note's row + VML shape survive verbatim alongside the new shadow
    f5 = os.path.join(tmp, "mixed.xlsx")
    _base(f5)
    _add_classic_note(f5)                     # D4, author "Deval"
    try:
        cid5 = xc.add_comment(f5, {"sheet": "Data", "cell": "B2"},
                              "Thread next to a note.", author="Alex Morgan")
    except _errors.CommentError as e:
        fails.append(f"coexist add: still refused: {e}")
        cid5 = None
    if cid5:
        body = _parts(f5)["xl/comments1.xml"].decode()
        for frag, what in (("<author>Deval</author>", "classic author"),
                           ('ref="D4"', "classic note row"),
                           ("Classic note text", "classic note text"),
                           ("tc=", "thread-shadow author"),
                           ('ref="B2"', "thread shadow row")):
            if frag not in body:
                fails.append(f"coexist add: {what} missing from comments1.xml")
        vml = _parts(f5)["xl/drawings/vmlDrawing1.vml"].decode()
        if "_x0000_s1025" not in vml:
            fails.append("coexist add: preserved note shape _x0000_s1025 missing from VML")
        if "_x0000_s1026" not in vml:
            fails.append("coexist add: new thread shape (id above preserved) missing from VML")
        if (o := _opens(f5)) is not True:
            fails.append(f"coexist add: won't open: {o}")

        # (b) the exact note-holding cell still refuses a thread (Excel's one-per-cell rule)
        try:
            xc.add_comment(f5, {"sheet": "Data", "cell": "D4"}, "x", author="A")
            fails.append("coexist per-cell: expected CommentError for the note's own cell")
        except _errors.CommentError as e:
            if "classic note" not in str(e):
                fails.append(f"coexist per-cell: message lacks 'classic note': {e}")

        # (c) reply keeps the classic content byte-identical in substance
        xc.reply(f5, cid5, "A reply.", author="Sam Lee")
        body = _parts(f5)["xl/comments1.xml"].decode()
        if "Classic note text" not in body or "<author>Deval</author>" not in body:
            fails.append("coexist reply: classic note lost on rebuild")

        # (d) deleting the LAST thread preserves the note layer (old: destroyed it)
        xc.delete(f5, cid5)                   # root -> takes its reply too
        parts5 = _parts(f5)
        if "xl/comments1.xml" not in parts5:
            fails.append("coexist teardown: comments1.xml destroyed with the last thread")
        else:
            body = parts5["xl/comments1.xml"].decode()
            if "Classic note text" not in body:
                fails.append("coexist teardown: classic note text lost")
            if "tc=" in body:
                fails.append("coexist teardown: thread shadow left behind")
        if "xl/drawings/vmlDrawing1.vml" not in parts5:
            fails.append("coexist teardown: VML destroyed with the last thread")
        elif "_x0000_s1025" not in parts5["xl/drawings/vmlDrawing1.vml"].decode():
            fails.append("coexist teardown: preserved note shape lost")
        if any(n.startswith("xl/threadedComments/") for n in parts5):
            fails.append("coexist teardown: threaded part not removed")
        if b"legacyDrawing" not in parts5["xl/worksheets/sheet1.xml"]:
            fails.append("coexist teardown: legacyDrawing removed from the sheet")
        if (o := _opens(f5)) is not True:
            fails.append(f"coexist teardown: won't open: {o}")

    # (review fix) a NON-canonically spelled classic ref still blocks its cell and
    # keeps its shape — 'd4' and 'D4' are the same cell
    f5b = os.path.join(tmp, "lowref.xlsx")
    _base(f5b)
    _add_classic_note(f5b, ref="d4")
    try:
        xc.add_comment(f5b, {"sheet": "Data", "cell": "D4"}, "x", author="A")
        fails.append("lowercase ref: expected CommentError for the note's cell")
    except _errors.CommentError as e:
        if "classic note" not in str(e):
            fails.append(f"lowercase ref: message wrong: {e}")
    xc.add_comment(f5b, {"sheet": "Data", "cell": "B2"}, "beside it", author="A")
    if "_x0000_s1025" not in _parts(f5b)["xl/drawings/vmlDrawing1.vml"].decode():
        fails.append("lowercase ref: classic shape lost on rebuild (canonical matching)")

    # (review fix) a Note shape WITHOUT <x:Row>/<x:Column> is preserved, not dropped
    f5c = os.path.join(tmp, "norowcol.xlsx")
    _base(f5c)
    _add_classic_note(f5c, with_row_col=False)
    xc.add_comment(f5c, {"sheet": "Data", "cell": "B2"}, "beside it", author="A")
    vml = _parts(f5c)["xl/drawings/vmlDrawing1.vml"].decode()
    if "_x0000_s1025" not in vml:
        fails.append("no-rowcol shape: preserved classic shape lost")
    if vml.count("<v:shape ") != 2:
        fails.append(f"no-rowcol shape: expected 2 shapes, got {vml.count('<v:shape ')}")

    # (e) form-control VML still refused (kept safety)
    f6 = os.path.join(tmp, "ctrl.xlsx")
    _base(f6)

    def add_checkbox(ps):
        shape = _NOTE_SHAPE.format(sid=1025, otype="Checkbox", rowcol=_rowcol(1, 1))
        ps.set_bytes("xl/drawings/vmlDrawing1.vml", _VML_NOTE.format(shapes=shape).encode("utf-8"))
        ps.ensure_default("vml", "application/vnd.openxmlformats-officedocument.vmlDrawing")
        ps.rels_for("xl/worksheets/sheet1.xml").add_rel(RNS + "/vmlDrawing",
                                                        "../drawings/vmlDrawing1.vml")
    patch_parts(f6, add_checkbox)
    try:
        xc.add_comment(f6, {"sheet": "Data", "cell": "B2"}, "x", author="A")
        fails.append("form-control: expected CommentError")
    except _errors.CommentError as e:
        if "form controls" not in str(e):
            fails.append(f"form-control: message changed: {e}")

    if fails:
        print("FAIL:")
        for x in fails:
            print("  -", x)
        return 1
    print("PASS — xlsx threaded comments (add / reply / resolve / reopen / delete)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
