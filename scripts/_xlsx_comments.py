"""Modern threaded comments for Excel (.xlsx) — list / add / reply / resolve / delete.

Excel keeps TWO synchronized representations and needs BOTH (confirmed against real Excel —
a threaded-only file has its comments dropped on open):

    xl/threadedComments/threadedComment{n}.xml   the modern thread (our source of truth)
    xl/persons/person.xml                        author registry
    xl/comments{n}.xml + xl/drawings/vmlDrawing{n}.vml + <legacyDrawing> in the sheet
                                                 the legacy "sticky note" that REGISTERS the
                                                 comment's existence on the cell

Threading is by `parentId` (reply -> root id); resolve is `done="1"` on the root threaded
comment. We treat the threaded part as authoritative and REGENERATE the legacy shadow from it
after every change, so the two never drift. Classic (non-threaded) NOTES already on the sheet
are PRESERVED through that regeneration (v0.3.0 coexistence) — their rows and VML shapes are
carried over verbatim; only the cell that holds a note refuses a new thread (Excel's own
one-per-cell rule).
"""
from __future__ import annotations

import copy
import re
import zipfile

from lxml import etree

from _ooxml_zip import patch_parts, _rels_name
from _errors import AnchorNotFound, CommentNotFound, CommentError
from _util import guid, utc_now

# --- namespaces ---
TC = "http://schemas.microsoft.com/office/spreadsheetml/2018/threadedcomments"
X = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
XR = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
MC = "http://schemas.openxmlformats.org/markup-compatibility/2006"
XML = "http://www.w3.org/XML/1998/namespace"

CT_THREADED = "application/vnd.ms-excel.threadedcomments+xml"
CT_PERSON = "application/vnd.ms-excel.person+xml"
CT_COMMENTS = "application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml"
CT_VML = "application/vnd.openxmlformats-officedocument.vmlDrawing"
REL_THREADED = "http://schemas.microsoft.com/office/2017/10/relationships/threadedComment"
REL_PERSON = "http://schemas.microsoft.com/office/2017/10/relationships/person"
REL_COMMENTS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/comments"
REL_VML = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/vmlDrawing"

PERSONS_PART = "xl/persons/person.xml"

_LEGACY_NOTE = ("[Threaded comment]\n\nYour version of Excel allows you to read this threaded "
                "comment; however, any edits to it will get removed if the file is opened in a "
                "newer version of Excel. Learn more: https://go.microsoft.com/fwlink/?linkid=870924")

_VML_OPEN = ('<xml xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office"'
             ' xmlns:x="urn:schemas-microsoft-com:office:excel">')
_VML_SHAPETYPE = ('<v:shapetype id="_x0000_t202" coordsize="21600,21600" o:spt="202" '
                  'path="m,l,21600r21600,l21600,xe"><v:stroke joinstyle="miter"/>'
                  '<v:path gradientshapeok="t" o:connecttype="rect"/></v:shapetype>')
_VML_TAIL = "</xml>"


def _vml_document(shapes, ids):
    """Assemble the VML note drawing. <o:idmap> must declare every 1024-wide shape-id block the
    notes actually span (ids start at _x0000_s1025), so the block list is computed from the
    ACTUAL ids present — preserved classic-note shapes plus our regenerated ones."""
    blocks = sorted({i // 1024 for i in ids}) or [1]
    idmap = ('<o:shapelayout v:ext="edit"><o:idmap v:ext="edit" '
             f'data="{",".join(map(str, blocks))}"/></o:shapelayout>')
    return _VML_OPEN + idmap + _VML_SHAPETYPE + "".join(shapes) + _VML_TAIL


# ---------------------------------------------------------------------------
# Cell-reference helpers
# ---------------------------------------------------------------------------

def _col_to_n(col):
    n = 0
    for ch in col.upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


MAX_COL, MAX_ROW = 16384, 1048576          # Excel's grid: columns A..XFD, rows 1..1048576


def _parse_ref(ref):
    m = re.match(r"([A-Za-z]{1,3})(\d+)$", str(ref).strip())
    if not m:
        raise AnchorNotFound(f"bad cell reference: {ref!r}")
    col0, row1 = _col_to_n(m.group(1)), int(m.group(2))
    if not (1 <= row1 <= MAX_ROW) or col0 >= MAX_COL:
        raise AnchorNotFound(f"cell reference out of range: {ref!r} "
                             f"(columns A..XFD, rows 1..{MAX_ROW})")
    return col0, row1 - 1   # (col0, row0)


def _ref_str(col0, row0):
    """Inverse of _parse_ref: canonical A1-style reference (so 'a01' round-trips to 'A1')."""
    n, s = col0 + 1, ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return f"{s}{row0 + 1}"


def _xl_dt(date):
    return date or (utc_now().strftime("%Y-%m-%dT%H:%M:%S") + ".00")


# ---------------------------------------------------------------------------
# Relationship / sheet resolution (driven by a get_bytes(name)->bytes|None callable)
# ---------------------------------------------------------------------------

def _rels_by_type(get_bytes, part):
    raw = get_bytes(_rels_name(part))
    by_id, by_type = {}, {}
    if raw:
        for rel in etree.fromstring(raw):
            by_id[rel.get("Id")] = rel.get("Target")
            by_type.setdefault(rel.get("Type"), []).append((rel.get("Id"), rel.get("Target")))
    return by_id, by_type


def _abs_target(base_part, target):
    if target.startswith("/"):       # absolute package path (openpyxl writes these)
        return target[1:]
    base = base_part.rsplit("/", 1)[0]
    out = []
    for p in (base + "/" + target).split("/"):
        if p == "..":
            if out:
                out.pop()
        elif p not in ("", "."):
            out.append(p)
    return "/".join(out)


def _sheet_map(get_bytes):
    wb = etree.fromstring(get_bytes("xl/workbook.xml"))
    by_id, _ = _rels_by_type(get_bytes, "xl/workbook.xml")
    out = {}
    for sh in wb.iter(f"{{{X}}}sheet"):
        tgt = by_id.get(sh.get(f"{{{R}}}id"))
        if tgt:
            out[sh.get("name")] = _abs_target("xl/workbook.xml", tgt)
    return out


def _sheet_part_rel(get_bytes, sheet_part, rel_type):
    _, by_type = _rels_by_type(get_bytes, sheet_part)
    rels = by_type.get(rel_type)
    return _abs_target(sheet_part, rels[0][1]) if rels else None


def _sheet_number(sheet_part):
    m = re.search(r"sheet(\d+)\.xml$", sheet_part)
    return m.group(1) if m else "1"


# ---------------------------------------------------------------------------
# Reading
# ---------------------------------------------------------------------------

def _persons_map(get_bytes):
    raw = get_bytes(PERSONS_PART)
    out = {}
    if raw:
        for p in etree.fromstring(raw).iter(f"{{{TC}}}person"):
            out[p.get("id")] = p.get("displayName")
    return out


def list_comments(path) -> list[dict]:
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
        get = lambda n: (z.read(n) if n in names else None)  # noqa: E731
        persons = _persons_map(get)
        sheet_map = _sheet_map(get)
        cell_sheets, wb = {}, None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            cell_sheets = {s: wb[s] for s in sheet_map if s in wb.sheetnames}
        except Exception:
            cell_sheets = {}

        records = []
        for sname, sheet_part in sheet_map.items():
            tpart = _sheet_part_rel(get, sheet_part, REL_THREADED)
            raw = get(tpart) if tpart else None
            if not raw:
                continue
            root = etree.fromstring(raw)
            tcs = list(root.iter(f"{{{TC}}}threadedComment"))
            done_of_root = {tc.get("id"): (tc.get("done") in ("1", "true"))
                            for tc in tcs if tc.get("parentId") is None}
            for tc in tcs:
                cid, parent, ref = tc.get("id"), tc.get("parentId"), tc.get("ref")
                root_id = parent or cid
                tnode = tc.find(f"{{{TC}}}text")
                val = None
                if sname in cell_sheets:
                    try:
                        val = cell_sheets[sname][ref].value
                    except Exception:
                        val = None
                records.append({
                    "id": cid,
                    "thread_id": root_id,
                    "parent_id": parent,
                    "is_reply": parent is not None,
                    "author": persons.get(tc.get("personId")),
                    "date": tc.get("dT"),
                    "text": tnode.text if tnode is not None else "",
                    "resolved": done_of_root.get(root_id, False),
                    "anchor": {"kind": "xlsx", "sheet": sname, "cell": ref},
                    "anchor_text": "" if val is None else str(val),
                    "context": "" if val is None else f"{ref} = {val!r}",
                    "location": f"{sname}!{ref}",
                })
        if wb is not None:
            try:
                wb.close()   # a lingering read handle would force writes off the atomic path
            except Exception:
                pass
    return records


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------

def _all_ids(ps):
    seen = set()
    if ps.has(PERSONS_PART):
        seen |= {p.get("id") for p in ps.get_xml(PERSONS_PART).iter(f"{{{TC}}}person")}
    for name in ps.names():
        if "threadedComment" in name and name.endswith(".xml"):
            seen |= {tc.get("id") for tc in ps.get_xml(name).iter(f"{{{TC}}}threadedComment")}
    return seen


def _ensure_person(ps, author):
    if ps.has(PERSONS_PART):
        root = ps.get_xml(PERSONS_PART)
        for p in root.findall(f"{{{TC}}}person"):
            if p.get("displayName") == author:
                return p.get("id")
    else:
        root = etree.Element(f"{{{TC}}}personList", nsmap={None: TC, "x": X})
        ps.add_xml_part(PERSONS_PART, root)
        ps.ensure_content_type("/" + PERSONS_PART, CT_PERSON)
        ps.rels_for("xl/workbook.xml").add_rel(REL_PERSON, "persons/person.xml")
    pid = guid(exclude=_all_ids(ps))
    p = etree.SubElement(root, f"{{{TC}}}person")
    p.set("displayName", author)
    p.set("id", pid)
    p.set("userId", author)
    p.set("providerId", "None")
    ps.set_xml(PERSONS_PART, root)
    return pid


def _threaded_for_sheet(ps, sheet_part):
    tpart = _sheet_part_rel(ps.get_bytes, sheet_part, REL_THREADED)
    if tpart and ps.has(tpart):
        return tpart, ps.get_xml(tpart)
    if not tpart:
        n = _sheet_number(sheet_part)
        tpart = f"xl/threadedComments/threadedComment{n}.xml"
        i = int(n)
        while ps.has(tpart):
            i += 1
            tpart = f"xl/threadedComments/threadedComment{i}.xml"
        ps.rels_for(sheet_part).add_rel(REL_THREADED, "../threadedComments/" + tpart.rsplit("/", 1)[1])
        ps.ensure_content_type("/" + tpart, CT_THREADED)
    root = etree.Element(f"{{{TC}}}ThreadedComments", nsmap={None: TC, "x": X})
    ps.add_xml_part(tpart, root)
    return tpart, root


def _thread_root_id(troot, cid):
    by_id = {tc.get("id"): tc for tc in troot.iter(f"{{{TC}}}threadedComment")}
    tc = by_id.get(cid)
    return (tc.get("parentId") or cid) if tc is not None else cid


def _find_comment(ps, cid):
    for sname, spart in _sheet_map(ps.get_bytes).items():
        tpart = _sheet_part_rel(ps.get_bytes, spart, REL_THREADED)
        if tpart and ps.has(tpart):
            for tc in ps.get_xml(tpart).iter(f"{{{TC}}}threadedComment"):
                if tc.get("id") == cid:
                    return sname, spart, tc.get("ref")
    return None


def _vml_shape(shape_id, col0, row0):
    anchor = f"{col0 + 1}, 15, {max(0, row0 - 1)}, 9, {col0 + 3}, 15, {row0 + 3}, 12"
    return (f'<v:shape id="_x0000_s{shape_id}" type="#_x0000_t202" '
            "style='position:absolute;margin-left:119.25pt;margin-top:6.75pt;width:108pt;"
            "height:59.25pt;z-index:1;visibility:hidden' fillcolor=\"infoBackground [80]\" "
            'strokecolor="none [81]" o:insetmode="auto">'
            '<v:fill color2="infoBackground [80]"/><v:shadow color="none [81]" obscured="t"/>'
            '<v:path o:connecttype="none"/>'
            "<v:textbox style='mso-direction-alt:auto'><div style='text-align:left'></div></v:textbox>"
            '<x:ClientData ObjectType="Note"><x:MoveWithCells/><x:SizeWithCells/>'
            f'<x:Anchor>{anchor}</x:Anchor>'
            f'<x:AutoFill>False</x:AutoFill><x:Row>{row0}</x:Row><x:Column>{col0}</x:Column>'
            '</x:ClientData></v:shape>')


def _legacy_thread_text(troot, root_id):
    lines = [_LEGACY_NOTE, ""]
    for tc in troot.iter(f"{{{TC}}}threadedComment"):
        if tc.get("id") == root_id or tc.get("parentId") == root_id:
            tnode = tc.find(f"{{{TC}}}text")
            lines.append("Comment:" if tc.get("parentId") is None else "Reply:")
            lines.append("    " + (tnode.text if tnode is not None else ""))
    return "\n".join(lines)


_V = "urn:schemas-microsoft-com:vml"
_XC = "urn:schemas-microsoft-com:office:excel"


def _check_vml_foreign_shapes(raw_v):
    """Refuse to regenerate the legacy VML when it holds NON-note drawings (form controls:
    checkboxes, buttons, dropdowns...) — they carry macro wiring and control parts this
    module can't safely re-emit, and Excel would silently drop mangled ones."""
    if raw_v:
        text = raw_v.decode("utf-8", "ignore")
        if len(re.findall(r"<v:shape\b", text)) > text.count('ObjectType="Note"'):
            raise CommentError("this sheet has non-comment drawings (e.g. form controls) in its "
                               "legacy VML; office-markup won't risk overwriting them")


def _canon_ref(raw):
    """Canonicalize a stored cell ref ('d4' -> 'D4') so comparisons against our own
    canonical refs can't miss; an unparseable ref is returned as-is."""
    try:
        c0, r0 = _parse_ref(raw or "")
        return _ref_str(c0, r0)
    except CommentError:
        return raw


def _classic_from_comments(raw_c):
    """Partition an existing xl/comments{n}.xml into the CLASSIC content to preserve.
    Returns (classic_authors, classic): the ordered, deduped author strings used by
    classic notes, and [(canonical ref, <comment> deep copy, author_str)] in original
    order. A comment is classic iff its author does NOT carry the `tc=` sentinel our
    thread shadows use. A corrupt/out-of-range authorId maps to '' (a synthesized author
    entry) rather than raising — never lose a note over a bad index."""
    if not raw_c:
        return [], []
    root = etree.fromstring(raw_c)
    authors = [a.text or "" for a in root.iter(f"{{{X}}}author")]
    seen_authors, out = [], []
    for c in root.iter(f"{{{X}}}comment"):
        try:
            a = authors[int(c.get("authorId") or 0)]
        except (ValueError, IndexError):
            a = ""
        if a.startswith("tc="):
            continue                                # our own thread shadow — regenerated
        if a not in seen_authors:
            seen_authors.append(a)
        out.append((_canon_ref(c.get("ref")), copy.deepcopy(c), a))
    return seen_authors, out


def _classic_vml(raw_v, classic_refs):
    """Extract, VERBATIM, the VML note shapes belonging to classic notes — matched by
    their ClientData Row/Column against the (canonical) classic refs. A Note shape whose
    position can't be read is preserved too: it cannot be one of OUR regenerated shapes
    (we always write Row/Column), and dropping it would desync note rows from shapes.
    Returns (shape_strings, ids)."""
    if not raw_v:
        return [], []
    want = set(classic_refs)
    shapes, ids = [], []
    for sh in etree.fromstring(raw_v).iter(f"{{{_V}}}shape"):
        cd = sh.find(f"{{{_XC}}}ClientData")
        if cd is None or cd.get("ObjectType") != "Note":
            continue
        try:
            keep = _ref_str(int(cd.findtext(f"{{{_XC}}}Column")),
                            int(cd.findtext(f"{{{_XC}}}Row"))) in want
        except (TypeError, ValueError):
            keep = True                             # unattributable -> preserve
        if keep:
            shapes.append(etree.tostring(sh, encoding="unicode"))
            m = re.search(r"_x0000_s(\d+)", sh.get("id") or "")
            if m:
                ids.append(int(m.group(1)))
    return shapes, ids


def _classic_note_refs(ps, sheet_part):
    """The cell refs on this sheet that hold a classic (non-threaded) note."""
    comments_part = _sheet_part_rel(ps.get_bytes, sheet_part, REL_COMMENTS)
    if not comments_part:
        return set()
    _, classic = _classic_from_comments(ps.get_bytes(comments_part))
    return {ref for ref, _c, _a in classic}


def _ensure_sheet_rel(ps, sheet_part, rel_type, target):
    """Return the Id of the sheet's relationship of `rel_type`, REUSING an existing one whatever its
    target spelling, and only adding `target` when none exists. Prevents a duplicate <Relationship>
    (which makes Excel drop the comments) when a foreign tool wrote the rel with a different spelling
    than our own `../...`."""
    _, by_type = _rels_by_type(ps.get_bytes, sheet_part)
    existing = by_type.get(rel_type)
    if existing:
        return existing[0][0]
    return ps.rels_for(sheet_part).add_rel(rel_type, target)


def _drop_sheet_rels(ps, sheet_part, rel_types):
    """Remove all of the sheet's relationships of the given types, whatever their target spelling —
    so a foreign `/xl/...`-spelled rel is unwired too, not left dangling to a dropped part."""
    _, by_type = _rels_by_type(ps.get_bytes, sheet_part)
    rels = ps.rels_for(sheet_part)
    for rt in rel_types:
        for _rid, tgt in by_type.get(rt, []):
            rels.remove_by_target(tgt)


def _remove_legacy_shadow(ps, sheet_part, comments_part, vml_part):
    """Tear the legacy shadow down when a sheet has no comments left: drop the comments + VML
    parts (and their content types), remove the <legacyDrawing>, and unwire the two rels."""
    ps.drop_part(comments_part)
    ps.drop_part(vml_part)
    sheet_root = ps.get_xml(sheet_part)
    ld = sheet_root.find(f"{{{X}}}legacyDrawing")
    if ld is not None:
        sheet_root.remove(ld)
        ps.set_xml(sheet_part, sheet_root)
    _drop_sheet_rels(ps, sheet_part, (REL_COMMENTS, REL_VML))


def _rebuild_legacy(ps, sheet_part, troot):
    """Regenerate xl/comments{n}.xml + xl/drawings/vmlDrawing{n}.vml from the current threads,
    PRESERVING any classic (non-threaded) notes already on the sheet: their <comment> rows and
    VML shapes are carried over verbatim (author ids remapped into the rebuilt list) and the
    thread shadows are rebuilt around them. The shadow is torn down only when the sheet has
    neither threads nor classic notes left."""
    n = _sheet_number(sheet_part)
    comments_part = _sheet_part_rel(ps.get_bytes, sheet_part, REL_COMMENTS) or f"xl/comments{n}.xml"
    vml_part = _sheet_part_rel(ps.get_bytes, sheet_part, REL_VML) or f"xl/drawings/vmlDrawing{n}.vml"
    roots = [tc for tc in troot.iter(f"{{{TC}}}threadedComment") if tc.get("parentId") is None]

    raw_c, raw_v = ps.get_bytes(comments_part), ps.get_bytes(vml_part)
    _check_vml_foreign_shapes(raw_v)
    classic_authors, classic = _classic_from_comments(raw_c)
    if not roots and not classic:
        _remove_legacy_shadow(ps, sheet_part, comments_part, vml_part)
        return

    croot = etree.Element(f"{{{X}}}comments", nsmap={None: X, "mc": MC, "xr": XR})
    croot.set(f"{{{MC}}}Ignorable", "xr")
    authors = etree.SubElement(croot, f"{{{X}}}authors")
    clist = etree.SubElement(croot, f"{{{X}}}commentList")

    # classic content first, in original order, with author ids remapped
    author_index = {}
    for a in classic_authors:
        author_index[a] = len(author_index)
        etree.SubElement(authors, f"{{{X}}}author").text = a
    for _ref, cel, a in classic:
        cel.set("authorId", str(author_index.get(a, 0)))
        clist.append(cel)
    shapes, ids = _classic_vml(raw_v, [ref for ref, _c, _a in classic])
    next_shape = max([1024] + ids) + 1        # new note shapes go ABOVE preserved ids

    for idx, rt in enumerate(roots):
        rid, ref = rt.get("id"), rt.get("ref")
        etree.SubElement(authors, f"{{{X}}}author").text = f"tc={rid}"
        c = etree.SubElement(clist, f"{{{X}}}comment")
        c.set("ref", ref)
        c.set("authorId", str(len(author_index) + idx))
        # Excel writes shapeId="0" on EVERY note (the cell comes from ref + the VML shape's
        # Row/Column, not from shapeId). A per-comment 1..N points at a VML shape id that does not
        # exist (real ids are 1025+), so Excel drops the whole comments part on open. See tests.
        c.set("shapeId", "0")
        c.set(f"{{{XR}}}uid", rid)
        t = etree.SubElement(etree.SubElement(c, f"{{{X}}}text"), f"{{{X}}}t")
        t.text = _legacy_thread_text(troot, rid)
        t.set(f"{{{XML}}}space", "preserve")
        col, row = _parse_ref(ref)
        shapes.append(_vml_shape(next_shape, col, row))
        ids.append(next_shape)
        next_shape += 1

    if ps.has(comments_part):
        ps.set_xml(comments_part, croot)
    else:
        ps.add_xml_part(comments_part, croot)
        ps.ensure_content_type("/" + comments_part, CT_COMMENTS)
    ps.set_bytes(vml_part, _vml_document(shapes, ids).encode("utf-8"))

    # wiring (idempotent): vml content-type default, the two sheet rels, and <legacyDrawing>
    ps.ensure_default("vml", CT_VML)
    _ensure_sheet_rel(ps, sheet_part, REL_COMMENTS, "../" + comments_part[len("xl/"):])
    vml_rid = _ensure_sheet_rel(ps, sheet_part, REL_VML, "../" + vml_part[len("xl/"):])
    sheet_root = ps.get_xml(sheet_part)
    ld = sheet_root.find(f"{{{X}}}legacyDrawing")
    if ld is None:
        ld = etree.Element(f"{{{X}}}legacyDrawing", nsmap={"r": R})   # bind r: so r:id (not ns0:id) serialises
        nxt = next((sheet_root.find(f"{{{X}}}{tag}") for tag in ("tableParts", "extLst")
                    if sheet_root.find(f"{{{X}}}{tag}") is not None), None)
        (nxt.addprevious if nxt is not None else sheet_root.append)(ld)
    ld.set(f"{{{R}}}id", vml_rid)
    ps.set_xml(sheet_part, sheet_root)


def _insert(ps, *, sheet, cell, text, author, date, parent_id):
    sheet_map = _sheet_map(ps.get_bytes)
    if parent_id is None:
        sheet = sheet or next(iter(sheet_map))
        if sheet not in sheet_map:
            raise AnchorNotFound(f"no sheet named {sheet!r}; have {list(sheet_map)}")
        sheet_part = sheet_map[sheet]
        col0, row0 = _parse_ref(cell)
        ref = _ref_str(col0, row0)             # canonical: 'a01' -> 'A1'
        # Excel allows ONE threaded comment thread per cell — refuse a duplicate root.
        existing = _sheet_part_rel(ps.get_bytes, sheet_part, REL_THREADED)
        if existing and ps.has(existing):
            for tc in ps.get_xml(existing).iter(f"{{{TC}}}threadedComment"):
                if tc.get("parentId") is None and tc.get("ref") == ref:
                    raise CommentError(f"cell {ref} already has a comment thread; reply to it instead")
        # ... and a cell holds EITHER a classic note OR a thread, never both (Excel's rule).
        # Other cells' notes are preserved; only this exact cell is refused.
        if ref in _classic_note_refs(ps, sheet_part):
            raise CommentError(f"cell {ref} already has a classic note; delete or convert the note "
                               f"in Excel first, or anchor the thread to another cell")
    else:
        loc = _find_comment(ps, parent_id)
        if loc is None:
            raise CommentNotFound(f"no comment with id {parent_id}")
        sheet, sheet_part, ref = loc

    pid = _ensure_person(ps, author)
    tpart, troot = _threaded_for_sheet(ps, sheet_part)
    new_id = guid(exclude=_all_ids(ps))
    tc = etree.SubElement(troot, f"{{{TC}}}threadedComment")
    tc.set("ref", ref)
    tc.set("dT", _xl_dt(date))
    tc.set("personId", pid)
    tc.set("id", new_id)
    if parent_id is not None:
        tc.set("parentId", _thread_root_id(troot, parent_id))
    etree.SubElement(tc, f"{{{TC}}}text").text = text

    ps.set_xml(tpart, troot)
    _rebuild_legacy(ps, sheet_part, troot)
    return new_id


def add_comment(path, anchor, text, *, author, initials=None, date=None) -> str:
    """Add a top-level comment. `anchor` = {"sheet": name|None, "cell": "B2"}."""
    out = {}
    patch_parts(path, lambda ps: out.__setitem__(
        "id", _insert(ps, sheet=anchor.get("sheet"), cell=anchor["cell"], text=text,
                      author=author, date=date, parent_id=None)))
    return out["id"]


def reply(path, parent_id, text, *, author, initials=None, date=None) -> str:
    out = {}
    patch_parts(path, lambda ps: out.__setitem__(
        "id", _insert(ps, sheet=None, cell=None, text=text, author=author, date=date,
                      parent_id=parent_id)))
    return out["id"]


def set_status(path, comment_id, resolved: bool) -> None:
    def mut(ps):
        loc = _find_comment(ps, comment_id)
        if loc is None:
            raise CommentNotFound(f"no comment with id {comment_id}")
        _, sheet_part, _ = loc
        tpart = _sheet_part_rel(ps.get_bytes, sheet_part, REL_THREADED)
        troot = ps.get_xml(tpart)
        root_id = _thread_root_id(troot, comment_id)
        for tc in troot.iter(f"{{{TC}}}threadedComment"):
            if tc.get("id") == root_id:
                tc.set("done", "1" if resolved else "0")
                ps.set_xml(tpart, troot)
                return
        raise CommentNotFound(f"no comment with id {comment_id}")
    patch_parts(path, mut)


def delete(path, comment_id) -> None:
    """Delete a comment. A reply goes alone; a thread root takes its whole thread."""
    def mut(ps):
        loc = _find_comment(ps, comment_id)
        if loc is None:
            raise CommentNotFound(f"no comment with id {comment_id}")
        _, sheet_part, _ = loc
        tpart = _sheet_part_rel(ps.get_bytes, sheet_part, REL_THREADED)
        troot = ps.get_xml(tpart)
        by_id = {tc.get("id"): tc for tc in troot.iter(f"{{{TC}}}threadedComment")}
        target = by_id[comment_id]
        if target.get("parentId") is None:
            doomed = [tc for tc in by_id.values()
                      if tc.get("id") == comment_id or tc.get("parentId") == comment_id]
        else:
            doomed = [target]
        for tc in doomed:
            troot.remove(tc)
        if troot.findall(f"{{{TC}}}threadedComment"):
            ps.set_xml(tpart, troot)
            _rebuild_legacy(ps, sheet_part, troot)
        else:
            # Last thread removed: drop the threaded part + its rel, then REBUILD the
            # legacy shadow — which preserves any classic notes on the sheet and only
            # tears the shadow down when nothing classic remains. (Pre-0.3.0 this path
            # removed the shadow unconditionally, destroying coexisting notes.)
            ps.drop_part(tpart)
            _drop_sheet_rels(ps, sheet_part, (REL_THREADED,))
            _rebuild_legacy(ps, sheet_part, etree.Element(f"{{{TC}}}ThreadedComments", nsmap={None: TC}))
    patch_parts(path, mut)
